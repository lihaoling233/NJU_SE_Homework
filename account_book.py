import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import datetime
from openpyxl import Workbook
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import os

# -------------------------- 数据库工具类 --------------------------
class DBUtil:
    def __init__(self, db_name="account_book.db"):
        self.db_name = db_name
        self.conn = None
        self.cursor = None
        self.init_db() #初始化数据库和表

    def connect(self):
        """建立数据库连接"""
        self.conn = sqlite3.connect(self.db_name)
        self.cursor = self.conn.cursor()

    def close(self):
        """关闭数据库连接"""
        if self.conn:
            self.conn.close()

    def init_db(self):
        """初始化数据库表和预定义分类"""
        self.connect()
        # 1. 用户表（单用户场景，简化设计）
        self.cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL DEFAULT '默认用户',
            create_time TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
        )
        ''')

        # 2. 分类表（收入/支出分类）
        self.cursor.execute('''
        CREATE TABLE IF NOT EXISTS categories (
            category_id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            type TEXT NOT NULL CHECK (type IN ('INCOME', 'EXPENSE')),
            is_default INTEGER NOT NULL DEFAULT 1  -- 1:预定义 0:自定义
        )
        ''')

        # 3. 交易表
        self.cursor.execute('''
        CREATE TABLE IF NOT EXISTS transactions (
            transaction_id INTEGER PRIMARY KEY AUTOINCREMENT,
            amount REAL NOT NULL,
            type TEXT NOT NULL CHECK (type IN ('INCOME', 'EXPENSE')),
            category_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            tag TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            create_time TEXT NOT NULL DEFAULT (datetime('now', 'localtime')),
            FOREIGN KEY (category_id) REFERENCES categories(category_id)
        )
        ''')

        # 4. 预算表（月度预算）
        self.cursor.execute('''
        CREATE TABLE IF NOT EXISTS budgets (
            budget_id INTEGER PRIMARY KEY AUTOINCREMENT,
            category_id INTEGER NOT NULL,
            month TEXT NOT NULL,  -- 格式：YYYY-MM
            amount REAL NOT NULL,
            spent REAL NOT NULL DEFAULT 0,
            FOREIGN KEY (category_id) REFERENCES categories(category_id),
            UNIQUE (category_id, month)  -- 同一分类同一月份唯一预算
        )
        ''')

        # 5. 设置表
        self.cursor.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            setting_id INTEGER PRIMARY KEY AUTOINCREMENT,
            remind_enabled INTEGER NOT NULL DEFAULT 0,  -- 0:关闭 1:开启
            remind_time TEXT DEFAULT '20:00',  -- 提醒时间
            user_id INTEGER NOT NULL DEFAULT 1,
            FOREIGN KEY (user_id) REFERENCES users(user_id)
        )
        ''')

        # 插入预定义分类（如果不存在）
        default_categories = [
            # 支出分类
            ('餐饮', 'EXPENSE', 1),
            ('交通', 'EXPENSE', 1),
            ('房租水电', 'EXPENSE', 1),
            ('购物', 'EXPENSE', 1),
            ('娱乐', 'EXPENSE', 1),
            # 收入分类
            ('工资', 'INCOME', 1),
            ('兼职', 'INCOME', 1),
            ('红包礼金', 'INCOME', 1),
            ('理财收益', 'INCOME', 1)
        ]
        for name, type_, is_default in default_categories:
            self.cursor.execute('''
            INSERT OR IGNORE INTO categories (name, type, is_default)
            VALUES (?, ?, ?)
            ''', (name, type_, is_default))

        # 初始化用户（如果不存在）
        self.cursor.execute('INSERT OR IGNORE INTO users (username) VALUES (?)', ('默认用户',))
        # 初始化设置（如果不存在）
        self.cursor.execute('INSERT OR IGNORE INTO settings (user_id) VALUES (?)', (1,))

        self.close()

    # -------------------------- 交易相关操作 --------------------------
    def add_transaction(self, amount, type_, category_id, date, tag='', remark=''):
        """添加交易记录（按月份同步预算，与天数无关）"""
        self.connect()
        try:
            # 1. 添加交易记录
            self.cursor.execute('''
            INSERT INTO transactions (amount, type, category_id, date, tag, remark)
            VALUES (?, ?, ?, ?, ?, ?)
            ''', (amount, type_, category_id, date, tag, remark))
            transaction_id = self.cursor.lastrowid

            # 2. 仅支出类型需要更新预算，且严格按月份匹配
            if type_ == 'EXPENSE':
                # 提取交易日期的月份（YYYY-MM），忽略具体天数
                transaction_month = date[:7]  # 如"2025-11-05" → "2025-11"
                
                # 查找该分类在该月份的预算
                self.cursor.execute('''
                SELECT budget_id, spent FROM budgets
                WHERE category_id = ? AND month = ?  -- 预算表的month是YYYY-MM，直接匹配
                ''', (category_id, transaction_month))
                budget = self.cursor.fetchone()
                
                if budget:
                    # 累加支出到该月预算的已花费中
                    budget_id, current_spent = budget
                    new_spent = current_spent + amount
                    self.cursor.execute('''
                    UPDATE budgets SET spent = ? WHERE budget_id = ?
                    ''', (new_spent, budget_id))

            self.conn.commit()
            return True, transaction_id
        except Exception as e:
            self.conn.rollback()
            return False, str(e)
        finally:
            self.close()

    def get_recent_transactions(self, limit=3):
        """获取最近N笔交易（含分类名称）"""
        self.connect()
        self.cursor.execute('''
        SELECT t.transaction_id, t.amount, t.type, c.name, t.date, t.tag, t.remark
        FROM transactions t
        JOIN categories c ON t.category_id = c.category_id
        ORDER BY t.date DESC, t.create_time DESC
        LIMIT ?
        ''', (limit,))
        result = self.cursor.fetchall()
        self.close()
        return result

    def get_transactions_by_condition(self, date=None, type_=None, category_id=None, tag=None, remark=None):
        """按条件查询交易记录（修复：支持年/年+月/精确日期查询）"""
        self.connect()
        query = '''
        SELECT t.transaction_id, t.amount, t.type, c.name, t.date, t.tag, t.remark
        FROM transactions t
        JOIN categories c ON t.category_id = c.category_id
        WHERE 1=1
        '''
        params = []

        # 修复：日期支持模糊匹配（年/年-月/精确日期）
        if date:
            date_len = len(date.strip())
            if date_len == 4:  # 仅年份（如2025）
                query += ' AND SUBSTR(t.date, 1, 4) = ?'  # 匹配前4位（年份）
                params.append(date.strip())
            elif date_len == 7:  # 年-月（如2025-10）
                query += ' AND SUBSTR(t.date, 1, 7) = ?'  # 匹配前7位（年-月）
                params.append(date.strip())
            elif date_len == 10:  # 精确日期（YYYY-MM-DD）
                query += ' AND t.date = ?'
                params.append(date.strip())

        if type_:
            query += ' AND t.type = ?'
            params.append(type_)
        if category_id:
            query += ' AND t.category_id = ?'
            params.append(category_id)
        if tag:
            query += ' AND t.tag LIKE ?'
            params.append(f'%{tag}%')
        if remark:
            query += ' AND t.remark LIKE ?'
            params.append(f'%{remark}%')

        query += ' ORDER BY t.date DESC, t.create_time DESC'  # 按账单日期排序（最新在前）
        self.cursor.execute(query, params)
        result = self.cursor.fetchall()
        self.close()
        return result

    def delete_duplicate_transactions(self):
        """删除重复交易（金额、类型、分类、日期、标签、备注完全相同）"""
        self.connect()
        try:
            # 找到重复记录的ID（保留最早一条）
            self.cursor.execute('''
            DELETE FROM transactions
            WHERE transaction_id NOT IN (
                SELECT MIN(transaction_id)
                FROM transactions
                GROUP BY amount, type, category_id, date, tag, remark
            )
            ''')
            deleted_count = self.cursor.rowcount
            self.conn.commit()
            return True, deleted_count
        except Exception as e:
            self.conn.rollback()
            return False, str(e)
        finally:
            self.close()

    # -------------------------- 分类相关操作 --------------------------
    def get_categories_by_type(self, type_):
        """按类型（INCOME/EXPENSE）获取分类"""
        self.connect()
        self.cursor.execute('''
        SELECT category_id, name FROM categories
        WHERE type = ?
        ORDER BY is_default DESC, name ASC
        ''', (type_,))
        result = self.cursor.fetchall()
        self.close()
        return result

    def add_custom_category(self, name, type_):
        """添加自定义分类"""
        self.connect()
        try:
            self.cursor.execute('''
            INSERT INTO categories (name, type, is_default)
            VALUES (?, ?, 0)
            ''', (name, type_))
            self.conn.commit()
            return True, self.cursor.lastrowid
        except Exception as e:
            self.conn.rollback()
            return False, str(e)
        finally:
            self.close()

    # -------------------------- 预算相关操作 --------------------------
    def set_monthly_budget(self, category_id, month, amount):
        """设置月度预算（严格按YYYY-MM匹配，与天数无关）"""
        self.connect()
        try:
            # 关键：仅按月份（YYYY-MM）汇总该分类的所有支出（忽略具体天数）
            self.cursor.execute('''
            SELECT COALESCE(SUM(amount), 0) FROM transactions
            WHERE category_id = ? 
            AND type = 'EXPENSE' 
            AND SUBSTR(date, 1, 7) = ?  -- 只取日期的前7位（YYYY-MM）进行匹配
            ''', (category_id, month))  # month必须是YYYY-MM格式（如"2025-11"）
            current_spent = self.cursor.fetchone()[0]  # 该月所有支出总和

            # 检查该分类该月份是否已有预算
            self.cursor.execute('''
            SELECT budget_id FROM budgets
            WHERE category_id = ? AND month = ?  -- 预算表的month是YYYY-MM，直接匹配
            ''', (category_id, month))
            budget = self.cursor.fetchone()

            if budget:
                # 已有预算：更新金额，同步该月最新支出总和
                budget_id = budget[0]
                self.cursor.execute('''
                UPDATE budgets 
                SET amount = ?, 
                    spent = ?  -- 强制同步该月所有支出（无论何时添加）
                WHERE budget_id = ?
                ''', (amount, current_spent, budget_id))
            else:
                # 新增预算：初始化已花费为该月所有支出总和
                self.cursor.execute('''
                INSERT INTO budgets (category_id, month, amount, spent)
                VALUES (?, ?, ?, ?)
                ''', (category_id, month, amount, current_spent))

            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            return False, str(e)
        finally:
            self.close()

    # -------------------------- 统计相关操作 --------------------------
    def get_monthly_statistics(self, month):
        """获取指定月份的收支统计（修复：确保正确匹配该月所有日期）"""
        self.connect()
        try:
            # 修复：用LIKE匹配该月份的所有日期（如'2025-10%'匹配2025-10-01至2025-10-31）
            self.cursor.execute('''
            SELECT COALESCE(SUM(amount), 0) FROM transactions
            WHERE type = 'INCOME' AND date LIKE ?
            ''', (f'{month}%',))  # 关键：添加通配符%
            total_income = self.cursor.fetchone()[0]

            self.cursor.execute('''
            SELECT COALESCE(SUM(amount), 0) FROM transactions
            WHERE type = 'EXPENSE' AND date LIKE ?
            ''', (f'{month}%',))  # 同理修复支出查询
            total_expense = self.cursor.fetchone()[0]

            balance = total_income - total_expense
            return total_income, total_expense, balance
        finally:
            self.close()

    def get_balance_trend(self, months=6):
        """获取近N个月的结余趋势（修复：基于账单实际存在的月份排序）"""
        self.connect()
        try:
            # 步骤1：从交易记录中提取所有不重复的月份（按账单日期）
            self.cursor.execute('''
            SELECT DISTINCT SUBSTR(date, 1, 7) AS month
            FROM transactions
            ORDER BY month DESC
            ''')
            all_months = [row[0] for row in self.cursor.fetchall()]  # 实际存在的月份列表

            # 步骤2：如果实际月份不足N个，用当前时间补全（避免空数据）
            current_month = datetime.datetime.now().strftime('%Y-%m')
            if not all_months:
                all_months = [current_month]
            # 补全缺失的月份（确保至少有N个）
            while len(all_months) < months:
                last_month = all_months[-1]
                last_date = datetime.datetime.strptime(last_month, '%Y-%m')
                prev_date = last_date - datetime.timedelta(days=1)  # 前一个月
                prev_month = prev_date.strftime('%Y-%m')
                all_months.append(prev_month)

            # 步骤3：取最近的N个月份，计算每个月的结余
            trend_data = []
            for month in all_months[:months]:  # 取前N个最新月份
                income, expense, balance = self.get_monthly_statistics(month)
                trend_data.append((month, balance))
            
            # 按月份正序排列（ oldest → newest ）
            trend_data.sort(key=lambda x: x[0])
            return trend_data
        finally:
            self.close()

    def get_monthly_expense_comparison(self):
        """获取本月与上月支出对比（本月支出、上月支出、变化率）"""
        self.connect()
        try:
            # 本月
            current_month = datetime.datetime.now().strftime('%Y-%m')
            _, current_expense, _ = self.get_monthly_statistics(current_month)

            # 上月
            last_month_date = datetime.datetime.now() - datetime.timedelta(days=30)
            last_month = last_month_date.strftime('%Y-%m')
            _, last_expense, _ = self.get_monthly_statistics(last_month)

            # 计算变化率（避免除零）
            if last_expense == 0:
                change_rate = 100.0 if current_expense > 0 else 0.0
            else:
                change_rate = ((current_expense - last_expense) / last_expense) * 100

            return current_expense, last_expense, change_rate
        finally:
            self.close()  # 所有操作完成后再关闭连接

    # -------------------------- 其他操作 --------------------------
    def export_transactions_to_excel(self, file_path):
        """导出所有交易记录到Excel"""
        try:
            # 获取所有交易
            transactions = self.get_transactions_by_condition()
            if not transactions:
                return False, "无交易数据可导出"

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "交易记录"

            # 写入表头
            headers = ['交易ID', '金额', '类型', '分类', '日期', '标签', '备注']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # 写入数据
            for row, trans in enumerate(transactions, 2):
                trans_id, amount, type_, category, date, tag, remark = trans
                # 转换类型显示（中文）
                type_cn = '收入' if type_ == 'INCOME' else '支出'
                ws.cell(row=row, column=1, value=trans_id)
                ws.cell(row=row, column=2, value=amount)
                ws.cell(row=row, column=3, value=type_cn)
                ws.cell(row=row, column=4, value=category)
                ws.cell(row=row, column=5, value=date)
                ws.cell(row=row, column=6, value=tag)
                ws.cell(row=row, column=7, value=remark)

            # 保存文件
            wb.save(file_path)
            return True, "导出成功"
        except Exception as e:
            return False, str(e)

    def update_remind_setting(self, enabled, time):
        """更新记账提醒设置"""
        self.connect()
        try:
            self.cursor.execute('''
            UPDATE settings SET remind_enabled = ?, remind_time = ?
            WHERE user_id = 1
            ''', (enabled, time))
            self.conn.commit()
            return True
        except Exception as e:
            self.conn.rollback()
            return False, str(e)
        finally:
            self.close()

    def get_monthly_budget_status(self, month):
        """强制使用字符串格式化传递参数，确保SQL正确执行"""
        self.connect()
        try:
            # 直接拼接SQL（仅临时用于排查，确认参数传递是否正确）
            query = f'''
            SELECT 
                c.name, 
                b.amount, 
                COALESCE(SUM(t.amount), 0) AS spent, 
                (b.amount - COALESCE(SUM(t.amount), 0)) AS remain
            FROM budgets b
            JOIN categories c ON b.category_id = c.category_id
            LEFT JOIN transactions t ON 
                b.category_id = t.category_id 
                AND t.type = 'EXPENSE' 
                AND SUBSTR(t.date, 1, 7) = b.month
            WHERE b.month = '{month}'  -- 直接拼接月份参数（仅测试用）
            GROUP BY c.name, b.amount
            ORDER BY c.type DESC, c.name ASC
            '''
            self.cursor.execute(query)
            result = self.cursor.fetchall()
            return result
        except Exception as e:
            print(f"预算查询错误：{e}")  # 打印错误信息
            return []
        finally:
            self.close()

# -------------------------- 统计工具类（封装统计逻辑） --------------------------
class StatisticsManager:
    def __init__(self, db_util):
        self.db_util = db_util

    def get_current_month_stat(self):
        """获取当月统计数据"""
        current_month = datetime.datetime.now().strftime('%Y-%m')
        total_income, total_expense, balance = self.db_util.get_monthly_statistics(current_month)
        return {
            'month': current_month,
            'total_income': round(total_income, 2),
            'total_expense': round(total_expense, 2),
            'balance': round(balance, 2)
        }

    def get_balance_trend(self, months=6):
        """获取近N个月结余趋势"""
        return self.db_util.get_balance_trend(months)

    def get_expense_comparison(self):
        """获取本月与上月支出对比"""
        current_expense, last_expense, change_rate = self.db_util.get_monthly_expense_comparison()
        return {
            'current_expense': round(current_expense, 2),
            'last_expense': round(last_expense, 2),
            'change_rate': round(change_rate, 1)  # 保留1位小数
        }


# -------------------------- GUI界面类 --------------------------
class AccountBookApp:
    def __init__(self, root):
        self.root = root
        self.root.title("记账本APP")
        self.root.geometry("800x600")  # 初始窗口大小
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.root.resizable(True, True)

        import warnings
        warnings.filterwarnings("ignore", category=UserWarning,  message="findfont: *")
        # 全局变量
        self.current_frame = None  # 当前显示的内容帧
        self.current_transaction_type = tk.StringVar(value='EXPENSE')  # 记收支时的类型（默认支出）

        from matplotlib.font_manager import FontProperties
        self.chinese_font = FontProperties(fname='/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc', size=10)
        plt.rcParams['font.family'] = ['WenQuanYi Zen Hei', 'sans-serif']
        plt.rcParams['axes.unicode_minus'] = False

        # 自定义Entry组件，默认用中文字体（辅助中文输入）
        from tkinter import Entry
        class ChineseEntry(Entry):
            def __init__(self, master=None, **kw):
                kw['font'] = kw.get('font', ('WenQuanYi Zen Hei', 12))  # 默认中文字体
                super().__init__(master, **kw)
        tk.Entry = ChineseEntry

        # 初始化数据库工具和统计工具
        self.db_util = DBUtil()
        self.stat_manager = StatisticsManager(self.db_util)

        # 初始化界面（底部导航 + 内容区）
        self.init_navigation()
        self.show_home_frame()  # 默认显示首页

    def on_close(self):
        """主窗口关闭时的资源释放逻辑"""
        # 1. 关闭所有Matplotlib图表（避免资源占用）
        import matplotlib.pyplot as plt
        plt.close('all')
        
        # 2. 关闭数据库连接（避免数据库锁死）
        self.db_util.close()
        
        # 3. 销毁Tkinter主窗口，终止主循环
        self.root.destroy()

    def init_navigation(self):
        """初始化底部导航栏"""
        # 导航帧
        nav_frame = tk.Frame(self.root, bg='#f0f0f0', height=50)
        nav_frame.pack(side=tk.BOTTOM, fill=tk.X)
        nav_frame.pack_propagate(False)  # 固定高度

        # 导航按钮样式
        btn_style = ttk.Style()
        btn_style.configure('Nav.TButton', font=('Arial', 12), padding=10)

        # 四个导航按钮
        self.home_btn = ttk.Button(nav_frame, text="首页", style='Nav.TButton', command=self.show_home_frame)
        self.add_btn = ttk.Button(nav_frame, text="记收支", style='Nav.TButton', command=self.show_add_transaction_frame)
        self.stat_btn = ttk.Button(nav_frame, text="统计", style='Nav.TButton', command=self.show_statistics_frame)
        self.my_btn = ttk.Button(nav_frame, text="我的", style='Nav.TButton', command=self.show_my_frame)

        # 按钮布局（均分宽度）
        self.home_btn.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.add_btn.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.stat_btn.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.my_btn.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    def switch_frame(self, new_frame):
        """切换内容帧（销毁旧帧，显示新帧）"""
        if self.current_frame:
            self.current_frame.destroy()
        self.current_frame = new_frame
        self.current_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

    # -------------------------- 1. 首页界面 --------------------------
    def show_home_frame(self):
        """显示首页帧"""
        home_frame = tk.Frame(self.root, bg='white')

        # 1. 顶部标题
        title_label = tk.Label(home_frame, text=f"记账本 - 首页", font=('Arial', 16, 'bold'), bg='white')
        title_label.pack(pady=10)

        # 2. 月度统计卡片（总收入、总支出、结余）
        stat_data = self.stat_manager.get_current_month_stat()
        stat_frame = tk.Frame(home_frame, bg='white')
        stat_frame.pack(pady=10, fill=tk.X, padx=20)

        # 卡片样式
        card_style = {'font': ('Arial', 14), 'bg': 'white', 'bd': 2, 'relief': tk.GROOVE, 'padx': 20, 'pady': 15}

        # 总收入卡片（绿色）
        income_card = tk.Label(stat_frame, text=f"本月总收入\n¥{stat_data['total_income']}", **card_style, fg='green')
        # 总支出卡片（红色）
        expense_card = tk.Label(stat_frame, text=f"本月总支出\n¥{stat_data['total_expense']}", **card_style, fg='red')
        # 结余卡片（蓝色）
        balance_card = tk.Label(stat_frame, text=f"本月结余\n¥{stat_data['balance']}", **card_style, fg='blue')

        # 卡片布局（均分）
        income_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        expense_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        balance_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)

        # 3. 最近交易记录
        recent_frame = tk.Frame(home_frame, bg='white')
        recent_frame.pack(pady=10, fill=tk.BOTH, expand=True, padx=20)

        # 标题
        recent_title = tk.Label(recent_frame, text="最近3笔交易", font=('Arial', 14, 'bold'), bg='white')
        recent_title.pack(anchor=tk.W, pady=5)

        # 交易列表（Treeview）
        columns = ('date', 'type', 'category', 'amount', 'remark')
        style = ttk.Style()
        style.configure('Treeview.Select', background='#4a86e8', foreground='white')
        tree = ttk.Treeview(recent_frame, columns=columns, show='headings')
        # 设置表头
        tree.heading('date', text='日期')
        tree.heading('type', text='类型')
        tree.heading('category', text='分类')
        tree.heading('amount', text='金额')
        tree.heading('remark', text='备注')
        # 设置列宽
        tree.column('date', width=100)
        tree.column('type', width=80)
        tree.column('category', width=100)
        tree.column('amount', width=100)
        tree.column('remark', width=300)

        # 填充数据
        recent_trans = self.db_util.get_recent_transactions(3)
        for trans in recent_trans:
            trans_id, amount, type_, category, date, tag, remark = trans
            type_cn = '收入' if type_ == 'INCOME' else '支出'
            amount_str = f"¥{amount}" if type_ == 'INCOME' else f"-¥{amount}"
            amount_color = 'green' if type_ == 'INCOME' else 'red'
            # 插入行并设置颜色
            item_id = tree.insert('', tk.END, values=(date, type_cn, category, amount_str, remark))
            tree.tag_configure(f'color_{item_id}', foreground=amount_color)
            tree.item(item_id, tags=(f'color_{item_id}',))

        # 滚动条
        scrollbar = ttk.Scrollbar(recent_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(fill=tk.BOTH, expand=True)

        # 切换到首页帧
        self.switch_frame(home_frame)

    # -------------------------- 2. 记收支界面 --------------------------
    def show_add_transaction_frame(self):
        """显示记收支帧"""
        add_frame = tk.Frame(self.root, bg='white')

        # 1. 标题
        title_label = tk.Label(add_frame, text="记收支", font=('Arial', 16, 'bold'), bg='white')
        title_label.pack(pady=10)

        # 2. 表单容器
        form_frame = tk.Frame(add_frame, bg='white')
        form_frame.pack(pady=10, fill=tk.X, padx=50)

        # 2.1 金额输入
        amount_frame = tk.Frame(form_frame, bg='white')
        amount_frame.pack(pady=8, fill=tk.X)
        amount_label = tk.Label(amount_frame, text="金额（¥）：", font=('Arial', 12), bg='white', width=10, anchor=tk.W)
        self.amount_entry = tk.Entry(amount_frame, font=('Arial', 12), width=30)
        amount_label.pack(side=tk.LEFT)
        self.amount_entry.pack(side=tk.LEFT)

        # 2.2 类型选择（收入/支出）
        type_frame = tk.Frame(form_frame, bg='white')
        type_frame.pack(pady=8, fill=tk.X)
        type_label = tk.Label(type_frame, text="类型：", font=('Arial', 12), bg='white', width=10, anchor=tk.W)
        expense_radio = tk.Radiobutton(type_frame, text="支出", variable=self.current_transaction_type, value='EXPENSE', font=('Arial', 12), bg='white')
        income_radio = tk.Radiobutton(type_frame, text="收入", variable=self.current_transaction_type, value='INCOME', font=('Arial', 12), bg='white')
        type_label.pack(side=tk.LEFT)
        expense_radio.pack(side=tk.LEFT, padx=10)
        income_radio.pack(side=tk.LEFT, padx=10)

        # 2.3 分类选择（根据类型动态加载）
        category_frame = tk.Frame(form_frame, bg='white')
        category_frame.pack(pady=8, fill=tk.X)
        category_label = tk.Label(category_frame, text="分类：", font=('Arial', 12), bg='white', width=10, anchor=tk.W)
        self.category_var = tk.StringVar()
        # 动态加载分类选项
        self.category_combobox = ttk.Combobox(category_frame, textvariable=self.category_var, font=('Arial', 12), width=28, state='readonly')
        self.load_categories()
        category_label.pack(side=tk.LEFT)
        self.category_combobox.pack(side=tk.LEFT)
        # 类型变化时重新加载分类
        self.current_transaction_type.trace('w', lambda *args: self.load_categories())

        # 2.4 日期选择（默认当天）
        date_frame = tk.Frame(form_frame, bg='white')
        date_frame.pack(pady=8, fill=tk.X)
        date_label = tk.Label(date_frame, text="日期：", font=('Arial', 12), bg='white', width=10, anchor=tk.W)
        self.date_entry = tk.Entry(date_frame, font=('Arial', 12), width=30)
        self.date_entry.insert(0, datetime.datetime.now().strftime('%Y-%m-%d'))
        date_label.pack(side=tk.LEFT)
        self.date_entry.pack(side=tk.LEFT)

        # 2.5 标签输入
        tag_frame = tk.Frame(form_frame, bg='white')
        tag_frame.pack(pady=8, fill=tk.X)
        tag_label = tk.Label(tag_frame, text="标签：", font=('Arial', 12), bg='white', width=10, anchor=tk.W)
        self.tag_entry = tk.Entry(tag_frame, font=('Arial', 12), width=30)
        tag_label.pack(side=tk.LEFT)
        self.tag_entry.pack(side=tk.LEFT)

        # 2.6 备注输入
        remark_frame = tk.Frame(form_frame, bg='white')
        remark_frame.pack(pady=8, fill=tk.X)
        remark_label = tk.Label(remark_frame, text="备注：", font=('Arial', 12), bg='white', width=10, anchor=tk.W)
        self.remark_entry = tk.Entry(remark_frame, font=('Arial', 12), width=30)
        remark_label.pack(side=tk.LEFT)
        self.remark_entry.pack(side=tk.LEFT)

        # 3. 保存按钮
        save_btn = ttk.Button(add_frame, text="保存交易", command=self.save_transaction, style='Accent.TButton')
        save_btn.pack(pady=20)

        # 切换到记收支帧
        self.switch_frame(add_frame)

    def load_categories(self):
        """根据当前选择的类型（收入/支出）加载分类到下拉框"""
        type_ = self.current_transaction_type.get()
        categories = self.db_util.get_categories_by_type(type_)  # [(id, name), ...]
        if not categories:
            self.category_combobox['values'] = []
            return
        # 提取分类名称和ID（用字典映射名称到ID）
        self.category_map = {name: id_ for id_, name in categories}
        self.category_combobox['values'] = [name for _, name in categories]
        # 默认选择第一个
        if categories:
            self.category_var.set(categories[0][1])

    def save_transaction(self):
        """保存交易记录（修复：新增支出后刷新预算状态）"""
        # 原有数据验证逻辑（保持不变）
        amount_str = self.amount_entry.get().strip()
        type_ = self.current_transaction_type.get()
        category_name = self.category_var.get()
        date = self.date_entry.get().strip()
        tag = self.tag_entry.get().strip()
        remark = self.remark_entry.get().strip()

        if not amount_str:
            messagebox.showerror("错误", "请输入金额！")
            return
        try:
            amount = float(amount_str)
            if amount <= 0:
                raise ValueError("金额必须大于0")
        except ValueError:
            messagebox.showerror("错误", "请输入有效的正数金额！")
            return

        if not category_name:
            messagebox.showerror("错误", "请选择分类！")
            return
        category_id = self.category_map.get(category_name)
        if not category_id:
            messagebox.showerror("错误", "所选分类无效！")
            return

        try:
            datetime.datetime.strptime(date, '%Y-%m-%d')
        except ValueError:
            messagebox.showerror("错误", "日期格式错误，请使用YYYY-MM-DD！")
            return

        # 保存交易
        success, result = self.db_util.add_transaction(amount, type_, category_id, date, tag, remark)
        if success:
            messagebox.showinfo("成功", "交易记录保存成功！")
            # 清空表单
            self.amount_entry.delete(0, tk.END)
            self.tag_entry.delete(0, tk.END)
            self.remark_entry.delete(0, tk.END)
            # 关键修复：如果是当月支出，刷新预算状态（如果当前在我的页面）
            current_month = datetime.datetime.now().strftime('%Y-%m')
            if type_ == 'EXPENSE' and date[:7] == current_month:
                # 检查当前是否在“我的页面”，如果是则刷新
                if hasattr(self.current_frame, 'winfo_children'):
                    children = self.current_frame.winfo_children()
                    if children and "我的" in children[0].cget("text"):
                        self.show_my_frame()
            # 刷新首页（如果当前在首页）
            elif isinstance(self.current_frame, tk.Frame) and "首页" in self.current_frame.winfo_children()[0].cget("text"):
                self.show_home_frame()
        else:
            messagebox.showerror("错误", f"保存失败：{result}")

    # -------------------------- 3. 统计界面 --------------------------
    def show_statistics_frame(self):
        """显示统计帧"""
        stat_frame = tk.Frame(self.root, bg='white')

        # 1. 标题
        title_label = tk.Label(stat_frame, text="统计分析", font=('Arial', 16, 'bold'), bg='white')
        title_label.pack(pady=10)

        # 2. 月度统计卡片（同首页，更详细）
        stat_data = self.stat_manager.get_current_month_stat()
        card_frame = tk.Frame(stat_frame, bg='white')
        card_frame.pack(pady=10, fill=tk.X, padx=20)

        # 卡片样式
        card_style = {'font': ('Arial', 14), 'bg': 'white', 'bd': 2, 'relief': tk.GROOVE, 'padx': 30, 'pady': 20}
        income_card = tk.Label(card_frame, text=f"本月总收入\n¥{stat_data['total_income']}", **card_style, fg='green')
        expense_card = tk.Label(card_frame, text=f"本月总支出\n¥{stat_data['total_expense']}", **card_style, fg='red')
        balance_card = tk.Label(card_frame, text=f"本月结余\n¥{stat_data['balance']}", **card_style, fg='blue')
        income_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        expense_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        balance_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)

        # 3. 月度支出对比
        comp_data = self.stat_manager.get_expense_comparison()
        comp_frame = tk.Frame(stat_frame, bg='white', bd=2, relief=tk.GROOVE)
        comp_frame.pack(pady=10, fill=tk.X, padx=20)
        comp_title = tk.Label(comp_frame, text="月度支出对比", font=('Arial', 12, 'bold'), bg='white')
        comp_title.pack(pady=5, anchor=tk.W, padx=10)
        comp_text = f"本月支出：¥{comp_data['current_expense']} | 上月支出：¥{comp_data['last_expense']} | 变化率：{comp_data['change_rate']}%"
        # 变化率颜色（负为减少，正为增加）
        comp_color = 'green' if comp_data['change_rate'] < 0 else 'red'
        comp_label = tk.Label(comp_frame, text=comp_text, font=('Arial', 12), bg='white', fg=comp_color)
        comp_label.pack(pady=5, padx=10)

        # 4. 结余趋势图（Matplotlib嵌入）
        trend_frame = tk.Frame(stat_frame, bg='white', bd=2, relief=tk.GROOVE)
        trend_frame.pack(pady=10, fill=tk.BOTH, expand=True, padx=20)
        trend_title = tk.Label(trend_frame, text="近6个月结余趋势", font=('Arial', 12, 'bold'), bg='white')
        trend_title.pack(pady=5, anchor=tk.W, padx=10)

        # 获取趋势数据
        trend_data = self.stat_manager.get_balance_trend(6)
        months = [item[0] for item in trend_data]
        balances = [item[1] for item in trend_data]

        # 创建图表
        #plt.rcParams['font.sans-serif'] = [' AR PL UMing CN']  # 支持中文
        #plt.rcParams['axes.unicode_minus'] = False  # 支持负号
        fig, ax = plt.subplots(figsize=(8, 4), dpi=100)
        ax.bar(months, balances, color=['blue' if b >= 0 else 'red' for b in balances])
        ax.set_xlabel('月份', fontproperties=self.chinese_font, fontsize=10)  # 横轴中文（🔶2-132需直观）
        ax.set_ylabel('结余金额（¥）', fontproperties=self.chinese_font, fontsize=10)  # 纵轴中文
        ax.set_title('结余趋势', fontproperties=self.chinese_font, fontsize=12)  # 标题中文（🔶2-132“分块清晰”）
        ax.grid(axis='y', linestyle='--', alpha=0.7)

        # 嵌入Tkinter
        canvas = FigureCanvasTkAgg(fig, master=trend_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 切换到统计帧
        self.switch_frame(stat_frame)

    # -------------------------- 4. 我的界面 --------------------------
    def show_my_frame(self):
        """显示我的帧"""
        my_frame = tk.Frame(self.root, bg='white')

        # 1. 标题
        title_label = tk.Label(my_frame, text="我的", font=('Arial', 16, 'bold'), bg='white')
        title_label.pack(pady=10)

        # 2. 功能按钮容器
        btn_frame = tk.Frame(my_frame, bg='white')
        btn_frame.pack(pady=20, fill=tk.X, padx=50)

        # 按钮样式
        btn_style = ttk.Style()
        btn_style.configure('My.TButton', font=('Arial', 12), padding=15)

        # 功能按钮
        budget_btn = ttk.Button(btn_frame, text="预算管理", style='My.TButton', command=self.show_budget_frame)
        history_btn = ttk.Button(btn_frame, text="历史交易记录", style='My.TButton', command=self.show_history_frame)
        data_btn = ttk.Button(btn_frame, text="数据管理", style='My.TButton', command=self.show_data_frame)
        setting_btn = ttk.Button(btn_frame, text="设置", style='My.TButton', command=self.show_setting_frame)

        # 按钮布局（两行两列）
        budget_btn.grid(row=0, column=0, padx=20, pady=10, sticky=tk.EW)
        history_btn.grid(row=0, column=1, padx=20, pady=10, sticky=tk.EW)
        data_btn.grid(row=1, column=0, padx=20, pady=10, sticky=tk.EW)
        setting_btn.grid(row=1, column=1, padx=20, pady=10, sticky=tk.EW)

        # 3. 当月预算状态（核心修复）
        budget_status_frame = tk.Frame(my_frame, bg='white', bd=2, relief=tk.GROOVE)
        budget_status_frame.pack(pady=10, fill=tk.BOTH, expand=True, padx=20)
        
        # 明确显示当前月份（YYYY-MM）
        current_month = datetime.datetime.now().strftime('%Y-%m')
        budget_status_title = tk.Label(
            budget_status_frame, 
            text=f"当月预算状态（{current_month}）",  # 标题显示当前月份
            font=('Arial', 12, 'bold'), 
            bg='white'
        )
        budget_status_title.pack(pady=5, anchor=tk.W, padx=10)

        # 预算表格（仅加载当前月份的预算）
        style = ttk.Style()
        style.configure('Treeview.Select', background='#4a86e8', foreground='white')
        budget_status = self.db_util.get_monthly_budget_status(current_month)  # 仅查当前月份
        columns = ('category', 'budget', 'spent', 'remain')
        tree = ttk.Treeview(budget_status_frame, columns=columns, show='headings')
        tree.heading('category', text='分类')
        tree.heading('budget', text='预算金额')
        tree.heading('spent', text='已花费（当月所有支出）')  # 明确说明是当月所有支出
        tree.heading('remain', text='剩余金额')
        tree.column('category', width=150)
        tree.column('budget', width=120)
        tree.column('spent', width=180)  # 加宽列，显示说明文字
        tree.column('remain', width=120)

        # 填充数据（剩余金额为负时标红）
        for status in budget_status:
            category, budget, spent, remain = status
            remain_color = 'red' if remain < 0 else 'black'
            item_id = tree.insert('', tk.END, values=(category, f"¥{budget}", f"¥{spent}", f"¥{remain}"))
            tree.tag_configure(f'remain_{item_id}', foreground=remain_color)
            tree.item(item_id, tags=(f'remain_{item_id}',))

        # 滚动条（保持不变）
        scrollbar = ttk.Scrollbar(budget_status_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 切换到我的帧
        self.switch_frame(my_frame)

    def show_budget_frame(self):
        """显示预算管理子帧"""
        budget_frame = tk.Frame(self.root, bg='white')

        # 标题
        title_label = tk.Label(budget_frame, text="预算管理", font=('Arial', 16, 'bold'), bg='white')
        title_label.pack(pady=10)

        # 表单容器
        form_frame = tk.Frame(budget_frame, bg='white')
        form_frame.pack(pady=10, fill=tk.X, padx=50)

        # 1. 月份选择（默认当月）
        month_frame = tk.Frame(form_frame, bg='white')
        month_frame.pack(pady=8, fill=tk.X)
        month_label = tk.Label(month_frame, text="月份：", font=('Arial', 12), bg='white', width=10, anchor=tk.W)
        self.budget_month_entry = tk.Entry(month_frame, font=('Arial', 12), width=30)
        default_month = datetime.datetime.now().strftime('%Y-%m')
        self.budget_month_entry.insert(0, default_month)
        month_label.pack(side=tk.LEFT)
        self.budget_month_entry.pack(side=tk.LEFT)

        # 2. 分类选择
        category_frame = tk.Frame(form_frame, bg='white')
        category_frame.pack(pady=8, fill=tk.X)
        category_label = tk.Label(category_frame, text="分类：", font=('Arial', 12), bg='white', width=10, anchor=tk.W)
        self.budget_category_var = tk.StringVar()
        # 加载所有支出分类（预算仅针对支出）
        expense_categories = self.db_util.get_categories_by_type('EXPENSE')
        self.budget_category_map = {name: id_ for id_, name in expense_categories}
        self.budget_category_combobox = ttk.Combobox(category_frame, textvariable=self.budget_category_var, font=('Arial', 12), width=28, state='readonly')
        self.budget_category_combobox['values'] = [name for _, name in expense_categories]
        if expense_categories:
            self.budget_category_var.set(expense_categories[0][1])
        category_label.pack(side=tk.LEFT)
        self.budget_category_combobox.pack(side=tk.LEFT)

        # 3. 预算金额
        amount_frame = tk.Frame(form_frame, bg='white')
        amount_frame.pack(pady=8, fill=tk.X)
        amount_label = tk.Label(amount_frame, text="预算金额（¥）：", font=('Arial', 12), bg='white', width=10, anchor=tk.W)
        self.budget_amount_entry = tk.Entry(amount_frame, font=('Arial', 12), width=30)
        amount_label.pack(side=tk.LEFT)
        self.budget_amount_entry.pack(side=tk.LEFT)

        # 4. 保存按钮
        save_btn = ttk.Button(budget_frame, text="保存预算", command=self.save_budget, style='Accent.TButton')
        save_btn.pack(pady=20)

        # 5. 返回按钮
        back_btn = ttk.Button(budget_frame, text="返回我的页面", command=self.show_my_frame)
        back_btn.pack(pady=10)

        # 切换到预算帧
        self.switch_frame(budget_frame)

    def save_budget(self):
        """保存预算（修复：保存后立即刷新预算状态显示）"""
        # 获取数据
        month = self.budget_month_entry.get().strip()
        category_name = self.budget_category_var.get()
        amount_str = self.budget_amount_entry.get().strip()

        # 数据验证（保持不变）
        if not month:
            messagebox.showerror("错误", "请输入月份！")
            return
        try:
            datetime.datetime.strptime(month, '%Y-%m')
        except ValueError:
            messagebox.showerror("错误", "月份格式错误，请使用YYYY-MM！")
            return

        if not category_name:
            messagebox.showerror("错误", "请选择分类！")
            return
        category_id = self.budget_category_map.get(category_name)
        if not category_id:
            messagebox.showerror("错误", "所选分类无效！")
            return

        if not amount_str:
            messagebox.showerror("错误", "请输入预算金额！")
            return
        try:
            amount = float(amount_str)
            if amount <= 0:
                raise ValueError("预算金额必须大于0")
        except ValueError:
            messagebox.showerror("错误", "请输入有效的正数预算金额！")
            return

        # 保存预算
        success = self.db_util.set_monthly_budget(category_id, month, amount)
        if success:
            messagebox.showinfo("成功", "预算设置保存成功！")
            # 关键修复：刷新预算状态显示（返回我的页面，重新加载预算列表）
            self.show_my_frame()
        else:
            messagebox.showerror("错误", "预算保存失败！")

    def show_history_frame(self):
        """显示历史交易记录子帧"""
        history_frame = tk.Frame(self.root, bg='white')

        # 标题
        title_label = tk.Label(history_frame, text="历史交易记录", font=('Arial', 16, 'bold'), bg='white')
        title_label.pack(pady=10)

        # 搜索表单
        search_frame = tk.Frame(history_frame, bg='white', bd=1, relief=tk.SUNKEN)
        search_frame.pack(pady=10, fill=tk.X, padx=20)

        # 搜索条件：日期、类型、分类、标签、备注
        # 1. 日期搜索
        date_frame = tk.Frame(search_frame, bg='white')
        date_frame.pack(pady=5, fill=tk.X, padx=10)
        date_label = tk.Label(date_frame, text="日期：", font=('Arial', 10), bg='white', width=8, anchor=tk.W)
        self.history_date_entry = tk.Entry(date_frame, font=('Arial', 10), width=20)
        date_label.pack(side=tk.LEFT)
        self.history_date_entry.pack(side=tk.LEFT, padx=5)

        # 2. 类型搜索
        type_frame = tk.Frame(search_frame, bg='white')
        type_frame.pack(pady=5, fill=tk.X, padx=10)
        type_label = tk.Label(type_frame, text="类型：", font=('Arial', 10), bg='white', width=8, anchor=tk.W)
        self.history_type_var = tk.StringVar(value='ALL')
        type_combobox = ttk.Combobox(type_frame, textvariable=self.history_type_var, font=('Arial', 10), width=18, state='readonly')
        type_combobox['values'] = ['全部', '收入', '支出']
        type_combobox.set('全部')
        type_label.pack(side=tk.LEFT)
        type_combobox.pack(side=tk.LEFT, padx=5)

        # 3. 分类搜索
        category_frame = tk.Frame(search_frame, bg='white')
        category_frame.pack(pady=5, fill=tk.X, padx=10)
        category_label = tk.Label(category_frame, text="分类：", font=('Arial', 10), bg='white', width=8, anchor=tk.W)
        self.history_category_var = tk.StringVar(value='ALL')
        # 加载所有分类
        all_categories = self.db_util.get_categories_by_type('INCOME') + self.db_util.get_categories_by_type('EXPENSE')
        self.history_category_map = {name: id_ for id_, name in all_categories}
        category_combobox = ttk.Combobox(category_frame, textvariable=self.history_category_var, font=('Arial', 10), width=18, state='readonly')
        category_combobox['values'] = ['全部'] + [name for _, name in all_categories]
        category_combobox.set('全部')
        category_label.pack(side=tk.LEFT)
        category_combobox.pack(side=tk.LEFT, padx=5)

        # 4. 标签和备注搜索
        tag_remark_frame = tk.Frame(search_frame, bg='white')
        tag_remark_frame.pack(pady=5, fill=tk.X, padx=10)
        tag_label = tk.Label(tag_remark_frame, text="标签：", font=('Arial', 10), bg='white', width=8, anchor=tk.W)
        self.history_tag_entry = tk.Entry(tag_remark_frame, font=('Arial', 10), width=18)
        remark_label = tk.Label(tag_remark_frame, text="备注：", font=('Arial', 10), bg='white', width=8, anchor=tk.W)
        self.history_remark_entry = tk.Entry(tag_remark_frame, font=('Arial', 10), width=18)
        tag_label.pack(side=tk.LEFT)
        self.history_tag_entry.pack(side=tk.LEFT, padx=5)
        remark_label.pack(side=tk.LEFT, padx=10)
        self.history_remark_entry.pack(side=tk.LEFT, padx=5)

        # 搜索按钮
        search_btn = ttk.Button(search_frame, text="搜索", command=self.search_history)
        search_btn.pack(pady=5)

        # 交易列表
        tree_frame = tk.Frame(history_frame, bg='white')
        tree_frame.pack(pady=10, fill=tk.BOTH, expand=True, padx=20)
        style = ttk.Style()
        style.configure('Treeview.Select', background='#4a86e8', foreground='white')
        columns = ('date', 'type', 'category', 'amount', 'tag', 'remark')
        self.history_tree = ttk.Treeview(tree_frame, columns=columns, show='headings')
        self.history_tree.heading('date', text='日期')
        self.history_tree.heading('type', text='类型')
        self.history_tree.heading('category', text='分类')
        self.history_tree.heading('amount', text='金额')
        self.history_tree.heading('tag', text='标签')
        self.history_tree.heading('remark', text='备注')
        self.history_tree.column('date', width=100)
        self.history_tree.column('type', width=80)
        self.history_tree.column('category', width=100)
        self.history_tree.column('amount', width=100)
        self.history_tree.column('tag', width=120)
        self.history_tree.column('remark', width=200)

        # 滚动条
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.history_tree.pack(fill=tk.BOTH, expand=True)

        # 加载所有历史记录
        self.search_history()

        # 返回按钮
        back_btn = ttk.Button(history_frame, text="返回我的页面", command=self.show_my_frame)
        back_btn.pack(pady=10)

        # 切换到历史帧
        self.switch_frame(history_frame)

    def search_history(self):
        """搜索历史交易记录"""
        # 获取搜索条件
        date = self.history_date_entry.get().strip() or None
        type_cn = self.history_type_var.get()
        type_ = None
        if type_cn == '收入':
            type_ = 'INCOME'
        elif type_cn == '支出':
            type_ = 'EXPENSE'

        category_name = self.history_category_var.get()
        category_id = None
        if category_name != '全部' and category_name in self.history_category_map:
            category_id = self.history_category_map[category_name]

        tag = self.history_tag_entry.get().strip() or None
        remark = self.history_remark_entry.get().strip() or None

        # 查询数据
        transactions = self.db_util.get_transactions_by_condition(date, type_, category_id, tag, remark)

        # 清空树
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)

        # 填充数据
        for trans in transactions:
            trans_id, amount, type_, category, date, tag, remark = trans
            type_cn = '收入' if type_ == 'INCOME' else '支出'
            amount_str = f"¥{amount}" if type_ == 'INCOME' else f"-¥{amount}"
            amount_color = 'green' if type_ == 'INCOME' else 'red'
            item_id = self.history_tree.insert('', tk.END, values=(date, type_cn, category, amount_str, tag, remark))
            self.history_tree.tag_configure(f'hist_color_{item_id}', foreground=amount_color)
            self.history_tree.item(item_id, tags=(f'hist_color_{item_id}',))

    def show_data_frame(self):
        """显示数据管理子帧"""
        data_frame = tk.Frame(self.root, bg='white')

        # 标题
        title_label = tk.Label(data_frame, text="数据管理", font=('Arial', 16, 'bold'), bg='white')
        title_label.pack(pady=10)

        # 功能按钮
        btn_frame = tk.Frame(data_frame, bg='white')
        btn_frame.pack(pady=20, fill=tk.X, padx=50)

        export_btn = ttk.Button(btn_frame, text="导出交易记录到Excel", command=self.export_excel, style='My.TButton')
        delete_dup_btn = ttk.Button(btn_frame, text="删除重复交易记录", command=self.delete_duplicates, style='My.TButton')
        export_btn.pack(pady=10, fill=tk.X)
        delete_dup_btn.pack(pady=10, fill=tk.X)

        # 数据统计信息
        stat_frame = tk.Frame(data_frame, bg='white', bd=2, relief=tk.GROOVE)
        stat_frame.pack(pady=20, fill=tk.X, padx=20)
        stat_title = tk.Label(stat_frame, text="数据统计", font=('Arial', 12, 'bold'), bg='white')
        stat_title.pack(pady=5, anchor=tk.W, padx=10)

        # 获取统计数据
        total_trans = len(self.db_util.get_transactions_by_condition())
        db_size = os.path.getsize('account_book.db') if os.path.exists('account_book.db') else 0
        stat_text = f"总交易记录数：{total_trans} | 数据库大小：{db_size} 字节"
        stat_label = tk.Label(stat_frame, text=stat_text, font=('Arial', 12), bg='white')
        stat_label.pack(pady=5, padx=10)

        # 返回按钮
        back_btn = ttk.Button(data_frame, text="返回我的页面", command=self.show_my_frame)
        back_btn.pack(pady=20)

        # 切换到数据帧
        self.switch_frame(data_frame)

    def export_excel(self):
        """导出Excel"""
        # 选择保存路径
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            title="选择Excel保存路径"
        )
        if not file_path:
            return

        # 导出
        success, msg = self.db_util.export_transactions_to_excel(file_path)
        if success:
            messagebox.showinfo("成功", msg)
        else:
            messagebox.showerror("错误", f"导出失败：{msg}")

    def delete_duplicates(self):
        """删除重复交易"""
        if not messagebox.askyesno("确认", "是否删除重复交易记录？（重复定义：金额、类型、分类、日期、标签、备注完全相同）"):
            return

        success, result = self.db_util.delete_duplicate_transactions()
        if success:
            messagebox.showinfo("成功", f"删除完成，共删除 {result} 条重复记录！")
            # 刷新历史记录（如果当前在历史页面）
            if isinstance(self.current_frame, tk.Frame) and "历史交易记录" in self.current_frame.winfo_children()[0].cget("text"):
                self.search_history()
        else:
            messagebox.showerror("错误", f"删除失败：{result}")

    def show_setting_frame(self):
        """显示设置子帧"""
        setting_frame = tk.Frame(self.root, bg='white')

        # 标题
        title_label = tk.Label(setting_frame, text="设置", font=('Arial', 16, 'bold'), bg='white')
        title_label.pack(pady=10)

        # 1. 记账提醒设置
        remind_frame = tk.Frame(setting_frame, bg='white', bd=1, relief=tk.SUNKEN)
        remind_frame.pack(pady=10, fill=tk.X, padx=20)
        remind_title = tk.Label(remind_frame, text="记账提醒", font=('Arial', 12, 'bold'), bg='white')
        remind_title.pack(pady=5, anchor=tk.W, padx=10)

        # 提醒开关
        self.remind_enabled_var = tk.IntVar(value=0)
        style = ttk.Style()
        style.configure('TCheckbutton', font=('Arial', 12), background='white')
        self.notification_var = tk.BooleanVar(value=True)
        remind_switch = ttk.Checkbutton(remind_frame, text="开启记账提醒", variable=self.remind_enabled_var)
        remind_switch.pack(pady=5, anchor=tk.W, padx=20)

        # 提醒时间
        time_frame = tk.Frame(remind_frame, bg='white')
        time_frame.pack(pady=5, anchor=tk.W, padx=20)
        time_label = tk.Label(time_frame, text="提醒时间：", font=('Arial', 10), bg='white')
        self.remind_time_entry = tk.Entry(time_frame, font=('Arial', 10), width=10)
        self.remind_time_entry.insert(0, '20:00')
        time_label.pack(side=tk.LEFT)
        self.remind_time_entry.pack(side=tk.LEFT)
        time_hint = tk.Label(time_frame, text="（格式：HH:MM）", font=('Arial', 8), bg='white', fg='gray')
        time_hint.pack(side=tk.LEFT, padx=5)

        # 2. 自定义分类设置
        custom_category_frame = tk.Frame(setting_frame, bg='white', bd=1, relief=tk.SUNKEN)
        custom_category_frame.pack(pady=10, fill=tk.X, padx=20)
        category_title = tk.Label(custom_category_frame, text="自定义分类", font=('Arial', 12, 'bold'), bg='white')
        category_title.pack(pady=5, anchor=tk.W, padx=10)

        # 分类表单
        form_frame = tk.Frame(custom_category_frame, bg='white')
        form_frame.pack(pady=5, anchor=tk.W, padx=20)
        name_label = tk.Label(form_frame, text="分类名称：", font=('Arial', 10), bg='white')
        self.custom_category_name = tk.Entry(form_frame, font=('Arial', 10), width=20)
        type_label = tk.Label(form_frame, text="类型：", font=('Arial', 10), bg='white')
        self.custom_category_type = tk.StringVar(value='EXPENSE')
        expense_radio = tk.Radiobutton(form_frame, text="支出", variable=self.custom_category_type, value='EXPENSE', font=('Arial', 10), bg='white')
        income_radio = tk.Radiobutton(form_frame, text="收入", variable=self.custom_category_type, value='INCOME', font=('Arial', 10), bg='white')

        name_label.pack(side=tk.LEFT)
        self.custom_category_name.pack(side=tk.LEFT, padx=5)
        type_label.pack(side=tk.LEFT, padx=10)
        expense_radio.pack(side=tk.LEFT, padx=5)
        income_radio.pack(side=tk.LEFT, padx=5)

        # 添加分类按钮
        add_category_btn = ttk.Button(custom_category_frame, text="添加自定义分类", command=self.add_custom_category, style='Small.TButton')
        add_category_btn.pack(pady=5, anchor=tk.W, padx=20)

        # 3. 保存设置按钮
        save_setting_btn = ttk.Button(setting_frame, text="保存设置", command=self.save_settings, style='Accent.TButton')
        save_setting_btn.pack(pady=20)

        # 4. 返回按钮
        back_btn = ttk.Button(setting_frame, text="返回我的页面", command=self.show_my_frame)
        back_btn.pack(pady=10)

        # 切换到设置帧
        self.switch_frame(setting_frame)

    def add_custom_category(self):
        """添加自定义分类"""
        name = self.custom_category_name.get().strip()
        type_ = self.custom_category_type.get()

        if not name:
            messagebox.showerror("错误", "请输入分类名称！")
            return

        # 检查分类是否已存在
        all_categories = self.db_util.get_categories_by_type('INCOME') + self.db_util.get_categories_by_type('EXPENSE')
        if name in [cat[1] for cat in all_categories]:
            messagebox.showerror("错误", "该分类名称已存在！")
            return

        # 添加
        success, result = self.db_util.add_custom_category(name, type_)
        if success:
            messagebox.showinfo("成功", "自定义分类添加成功！")
            self.custom_category_name.delete(0, tk.END)
            # 刷新分类下拉框（如果需要）
            if hasattr(self, 'category_combobox'):
                self.load_categories()
        else:
            messagebox.showerror("错误", f"添加失败：{result}")

    def save_settings(self):
        """保存设置"""
        # 获取提醒设置
        remind_enabled = self.remind_enabled_var.get()
        remind_time = self.remind_time_entry.get().strip()

        # 验证时间格式
        try:
            datetime.datetime.strptime(remind_time, '%H:%M')
        except ValueError:
            messagebox.showerror("错误", "提醒时间格式错误，请使用HH:MM！")
            return

        # 保存
        success = self.db_util.update_remind_setting(remind_enabled, remind_time)
        if success:
            messagebox.showinfo("成功", "设置保存成功！")
        else:
            messagebox.showerror("错误", "设置保存失败！")


# -------------------------- 程序入口 --------------------------
if __name__ == "__main__":
    root = tk.Tk()
    app = AccountBookApp(root)
    root.mainloop()
